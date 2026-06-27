import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    PayrollRun, PayrollRecord, Payslip, PayslipTemplate,
    SalaryStructure, SalaryComponent, EmployeeSalary,
    StatutorySetting, EmployeeLoan, ExpenseClaim,
    TaxDeclaration, Form16,
)
from utils.pdf import render_pdf_response
from utils.excel import make_workbook, apply_header_row, auto_fit_columns, workbook_response
from utils.access import manager_or_hr_required, employee_profile_required, can_manage_employee, perm_required, has_full_team_scope
from apps.tenants.jurisdiction import require_india_payroll, require_gcc_payroll, is_gcc_payroll, normalise_jurisdiction


# ---------------------------------------------------------------------------
# Payroll run management (HR admin)
# ---------------------------------------------------------------------------
@perm_required("payroll.view_all")
def payroll_run_list(request):
    tenant = request.tenant
    runs = PayrollRun.objects.filter(tenant=tenant).order_by("-year", "-month")
    return render(request, "payroll/run_list.html", {"runs": runs})


@perm_required("payroll.run")
def payroll_run_create(request):
    tenant = request.tenant
    if request.method == "POST":
        year = int(request.POST.get("year"))
        month = int(request.POST.get("month"))

        if PayrollRun.objects.filter(tenant=tenant, year=year, month=month).exists():
            messages.error(request, f"Payroll for {year}-{month:02d} already exists.")
            return redirect("payroll:run_list")

        run = PayrollRun.objects.create(
            tenant=tenant, year=year, month=month,
            status="draft", run_by=request.user
        )
        from .dispatch import dispatch_payroll_run
        mode = dispatch_payroll_run(tenant, run)
        if mode == "sync":
            messages.success(request, f"Payroll computed for {year}-{month:02d}. Review the run.")
        else:
            messages.success(request, f"Payroll run initiated for {year}-{month:02d}. Processing in background.")
        return redirect("payroll:run_detail", pk=run.pk)

    today = timezone.localdate()
    return render(request, "payroll/run_create.html", {"today": today})


@perm_required("payroll.view_all")
def payroll_run_detail(request, pk):
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    records = run.records.select_related(
        "employee", "employee__department", "employee__designation"
    ).order_by("employee__first_name")
    # Run anomaly detection — pure local rules, no external API
    from .anomaly import detect_for_run, summary
    anomalies = detect_for_run(run)
    from django.conf import settings
    finance_sync_available = bool(
        getattr(settings, "FIN_INTERNAL_BASE_URL", "")
        and getattr(settings, "SSO_SHARED_SECRET", "")
    )
    gcc_export_readiness = None
    if is_gcc_payroll(tenant):
        from .gcc_export_validation import assess_gcc_export_readiness
        gcc_export_readiness = assess_gcc_export_readiness(tenant, run)
    return render(request, "payroll/run_detail.html", {
        "run": run, "records": records,
        "anomalies": anomalies, "anomaly_counts": summary(anomalies),
        "finance_sync_available": finance_sync_available,
        "gcc_export_readiness": gcc_export_readiness,
    })


@perm_required("payroll.approve")
@require_POST
def payroll_run_approve(request, pk):
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    if run.status != "review":
        messages.error(request, "Only runs in 'review' status can be approved.")
        return redirect("payroll:run_detail", pk=pk)

    run.status = "approved"
    run.approved_by = request.user
    run.approved_at = timezone.now()
    run.save(update_fields=["status", "approved_by", "approved_at"])

    # Lock all records and generate payslips
    run.records.all().update(is_locked=True, locked_at=timezone.now())
    from .dispatch import dispatch_payslip_generation
    dispatch_payslip_generation(run)

    messages.success(request, "Payroll approved and payslips are being generated.")
    return redirect("payroll:run_detail", pk=pk)


@perm_required("payroll.run")
@require_POST
def payroll_run_publish(request, pk):
    """Publish payslips so employees can see them in ESS."""
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    if run.status != "approved":
        messages.error(request, "Only approved runs can be published.")
        return redirect("payroll:run_detail", pk=pk)

    payslips_to_publish = list(
        Payslip.objects.filter(
            tenant=tenant, year=run.year, month=run.month, is_published=False
        ).select_related("employee__user", "payroll_record")
    )
    Payslip.objects.filter(
        tenant=tenant, year=run.year, month=run.month, is_published=False,
    ).update(is_published=True, published_at=timezone.now())

    # Notify every employee whose payslip just went live
    try:
        from apps.hr_ops.services import notify
        from utils.money import format_money
        for slip in payslips_to_publish:
            if slip.employee and slip.employee.user:
                net_fmt = format_money(slip.payroll_record.net_payable, tenant.currency)
                notify(
                    slip.employee.user, "payslip_published",
                    f"Your payslip for {run.year}-{run.month:02d} is ready",
                    message=f"Net payable: {net_fmt}\n\nView the breakdown and download the PDF from your dashboard.",
                    action_url="/payroll/my-payslips/",
                )
    except Exception:
        pass

    # Audit log entry
    try:
        from apps.hr_ops.services import audit_log
        audit_log(
            tenant, request.user, "publish", "PayrollRun", run,
            f"Published {len(payslips_to_publish)} payslip(s) for {run.year}-{run.month:02d}",
            details={"count": len(payslips_to_publish), "total_net": float(run.total_net)},
        )
    except Exception:
        pass

    # Email PDF + WhatsApp summary (async when Celery is available)
    try:
        from apps.hr_ops.tasks import deliver_payslips_task
        deliver_payslips_task.delay(run.id)
    except Exception:
        try:
            from apps.hr_ops.payslip_delivery import deliver_payslips_for_run
            deliver_payslips_for_run(run)
        except Exception:
            pass

    # HR admin monthly summary email
    try:
        from apps.hr_ops.hr_report_email import notify_hr_payroll_published
        notify_hr_payroll_published(run, payslip_count=len(payslips_to_publish))
    except Exception:
        pass

    # Sync payroll journal to Finance (best-effort)
    finance_msg = ""
    try:
        finance_msg = _apply_finance_sync(tenant, run, request)
    except Exception:
        pass

    messages.success(
        request,
        f"Published {len(payslips_to_publish)} payslip(s). Email and WhatsApp delivery started.{finance_msg}",
    )
    return redirect("payroll:run_detail", pk=pk)


def _apply_finance_sync(tenant, run, request=None) -> str:
    """Post payroll journal to Finance; returns user-facing suffix message."""
    from .finance_sync import sync_payroll_to_finance

    if run.finance_synced_at:
        return " Finance ledger already synced."

    result = sync_payroll_to_finance(tenant, run)
    if not result.get("ok"):
        if request:
            messages.warning(request, f"Finance sync skipped: {result.get('error', 'unknown error')}")
        return ""

    run.finance_journal_id = str(result.get("journal_entry_id") or "")
    run.finance_voucher_no = result.get("voucher_no") or ""
    run.finance_synced_at = timezone.now()
    run.save(update_fields=["finance_journal_id", "finance_voucher_no", "finance_synced_at"])

    suffix = f" Finance journal {run.finance_voucher_no} posted."
    if request:
        messages.info(request, f"Payroll synced to Finance as voucher {run.finance_voucher_no}.")
    return suffix


@perm_required("payroll.run")
@require_POST
def payroll_sync_finance(request, pk):
    """Manually sync an approved/paid payroll run to Finance ledger."""
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    if run.status not in ("approved", "paid"):
        messages.error(request, "Only approved or paid runs can be synced to Finance.")
        return redirect("payroll:run_detail", pk=pk)

    if run.finance_synced_at:
        messages.info(request, f"Already synced as {run.finance_voucher_no or run.finance_journal_id}.")
        return redirect("payroll:run_detail", pk=pk)

    _apply_finance_sync(tenant, run, request)
    return redirect("payroll:run_detail", pk=pk)


# ---------------------------------------------------------------------------
# Pre-payroll review & per-employee adjustments
# ---------------------------------------------------------------------------
@perm_required("payroll.view_all")
def payroll_monthly_review(request):
    tenant = request.tenant
    today = timezone.localdate()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    from .review_services import build_monthly_readiness, prepare_month_attendance

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "prepare":
            count = prepare_month_attendance(tenant, year, month)
            messages.success(request, f"Attendance summaries computed for {count} employee(s).")
        elif action == "run_payroll":
            if PayrollRun.objects.filter(tenant=tenant, year=year, month=month).exists():
                messages.error(request, f"Payroll for {year}-{month:02d} already exists.")
            else:
                prepare_month_attendance(tenant, year, month)
                run = PayrollRun.objects.create(
                    tenant=tenant, year=year, month=month,
                    status="draft", run_by=request.user,
                )
                from .dispatch import dispatch_payroll_run
                mode = dispatch_payroll_run(tenant, run)
                if mode == "sync":
                    messages.success(request, "Payroll computed. Review the run.")
                else:
                    messages.info(request, "Payroll run started in background.")
                messages.success(request, f"Payroll run started for {year}-{month:02d}.")
                return redirect("payroll:run_detail", pk=run.pk)
        return redirect(f"{request.path}?year={year}&month={month}")

    data = build_monthly_readiness(tenant, year, month)
    return render(request, "payroll/monthly_review.html", {
        "year": year,
        "month": month,
        **data,
    })


@perm_required("payroll.view_all")
def payroll_record_detail(request, run_pk, record_pk):
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=run_pk, tenant=tenant)
    record = get_object_or_404(
        PayrollRecord, pk=record_pk, payroll_run=run, tenant=tenant,
    )
    from .review_services import get_employee_month_context

    ctx = get_employee_month_context(
        tenant, record.employee, run.year, run.month, payroll_run=run,
    )
    return render(request, "payroll/record_detail.html", {
        "run": run,
        "record": record,
        "employee": record.employee,
        **ctx,
    })


@perm_required("payroll.run")
@require_POST
def payroll_record_edit(request, run_pk, record_pk):
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=run_pk, tenant=tenant)
    if run.status not in ("review", "draft", "processing"):
        messages.error(request, "This payroll run can no longer be edited.")
        return redirect("payroll:run_detail", pk=run.pk)

    record = get_object_or_404(
        PayrollRecord, pk=record_pk, payroll_run=run, tenant=tenant,
    )
    if record.is_locked:
        messages.error(request, "This record is locked.")
        return redirect("payroll:record_detail", run_pk=run.pk, record_pk=record.pk)

    from decimal import Decimal, InvalidOperation

    lop_raw = (request.POST.get("lop_override") or "").strip()
    bonus_raw = (request.POST.get("bonus_amount") or "0").strip()
    ded_raw = (request.POST.get("manual_deduction") or "0").strip()
    hr_notes = (request.POST.get("hr_notes") or "").strip()

    try:
        record.lop_override = Decimal(lop_raw) if lop_raw else None
        record.bonus_amount = Decimal(bonus_raw or "0")
        record.manual_deduction = Decimal(ded_raw or "0")
        record.hr_notes = hr_notes
        record.save(update_fields=["lop_override", "bonus_amount", "manual_deduction", "hr_notes"])

        from .engine import compute_payroll_record
        from .review_services import refresh_run_totals

        compute_payroll_record(tenant, record.employee, run, run.year, run.month)
        refresh_run_totals(run)
        messages.success(request, f"Updated payroll for {record.employee.full_name} and recalculated.")
    except (InvalidOperation, ValueError) as exc:
        messages.error(request, f"Invalid values: {exc}")

    return redirect("payroll:record_detail", run_pk=run.pk, record_pk=record.pk)


@perm_required("payroll.run")
@require_POST
def payroll_run_recompute(request, pk):
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    if run.status not in ("review", "draft"):
        messages.error(request, "Only draft or review runs can be recomputed.")
        return redirect("payroll:run_detail", pk=pk)

    from .review_services import prepare_month_attendance, refresh_run_totals
    from .engine import compute_payroll_record

    prepare_month_attendance(tenant, run.year, run.month)
    for rec in run.records.select_related("employee"):
        try:
            compute_payroll_record(tenant, rec.employee, run, run.year, run.month)
        except ValueError as exc:
            messages.warning(request, str(exc))
    refresh_run_totals(run)
    messages.success(request, "Payroll recomputed from latest attendance.")
    return redirect("payroll:run_detail", pk=pk)


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------
def _earning_amount(record, code: str, fallback_field: str = "") -> float:
    """Read a component amount from denormalized field or earnings_detail JSON."""
    if fallback_field:
        val = getattr(record, fallback_field, None)
        if val and float(val) > 0:
            return float(val)
    detail = (record.earnings_detail or {}).get(code) or {}
    return float(detail.get("amount") or 0)


def _salary_register_gcc_excel(tenant, run, records):
    """GCC salary register — IBAN/SWIFT bank columns, no India statutory fields."""
    currency = tenant.currency or "AED"
    wb = make_workbook()
    ws = wb.active
    ws.title = f"Salary Register {run.year}-{run.month:02d}"

    headers = [
        "Emp Code", "Name", "Department", "Designation",
        "Bank Name", "Account Holder", "IBAN", "SWIFT/BIC", "Branch", "A/C Type",
        "Working Days", "LOP Days", "Paid Days",
        "Basic", "Housing", "Transport", "Other Earnings", f"Gross ({currency})",
        "Statutory (Emp)", "Loan", "Other Deductions", f"Total Deductions ({currency})",
        f"Net Payable ({currency})", "Statutory (Employer)", "EOS/Indemnity Accrual",
        f"Total Employer Cost ({currency})",
    ]
    apply_header_row(ws, headers)

    for rec in records:
        emp = rec.employee
        bank = emp.bank_accounts.filter(is_primary=True).first()
        statutory_emp = float(rec.pf_employee or 0)
        statutory_er = float(rec.pf_employer or 0)
        eos_accrual = _earning_amount(rec, "INDEM_ACCR") or _earning_amount(rec, "GRAT_ACCR")
        employer_cost = float(rec.gross_earnings or 0) + statutory_er + eos_accrual
        ws.append([
            emp.employee_code,
            emp.full_name,
            emp.department.name if emp.department else "",
            emp.designation.name if emp.designation else "",
            bank.bank_name if bank else "",
            bank.account_holder_name if bank else "",
            getattr(bank, "iban", "") or (bank.account_number if bank else ""),
            getattr(bank, "swift_code", "") or getattr(bank, "ifsc_code", "") or "",
            bank.branch_name if bank else "",
            bank.account_type if bank else "",
            rec.working_days,
            float(rec.lop_days),
            float(rec.paid_days),
            _earning_amount(rec, "BASIC", "basic"),
            _earning_amount(rec, "HOUSING", "hra"),
            _earning_amount(rec, "TRANSPORT", "conveyance"),
            _earning_amount(rec, "SPECIAL", "special_allowance") + float(rec.other_earnings or 0),
            float(rec.gross_earnings),
            statutory_emp,
            float(rec.loan_deduction),
            float(rec.other_deductions),
            float(rec.total_deductions),
            float(rec.net_payable),
            statutory_er,
            eos_accrual,
            employer_cost,
        ])

    auto_fit_columns(ws)
    return workbook_response(wb, f"salary_register_{run.year}_{run.month:02d}.xlsx")


@perm_required("payroll.export")
def salary_register_excel(request, pk):
    """
    Salary register Excel — primary document for bank disbursement and CA/compliance.
    India: IFSC + PF/ESI/PT/TDS columns. GCC: IBAN/SWIFT + statutory/EOS columns.
    """
    from apps.tenants.jurisdiction import is_gcc_payroll

    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    records = (
        run.records
        .select_related("employee", "employee__department", "employee__designation")
        .prefetch_related("employee__bank_accounts")
        .order_by("employee__employee_code")
    )
    if is_gcc_payroll(tenant.payroll_jurisdiction):
        return _salary_register_gcc_excel(tenant, run, records)

    wb = make_workbook()
    ws = wb.active
    ws.title = f"Salary Register {run.year}-{run.month:02d}"

    headers = [
        # Identity
        "Emp Code", "Name", "Department", "Designation",
        # Bank (for disbursement)
        "Bank Name", "Account Holder", "Account Number", "IFSC", "Branch", "A/C Type",
        # Attendance
        "Working Days", "LOP Days", "Paid Days",
        # Earnings
        "Basic", "HRA", "Conveyance", "Special Allowance", "Other Earnings", "Gross Earnings",
        # Statutory deductions
        "PF (Emp)", "ESI (Emp)", "Professional Tax", "LWF (Emp)", "TDS",
        # Other deductions
        "Loan EMI", "Other Deductions", "Total Deductions",
        # Net
        "Net Payable",
        # Employer cost (for finance reporting)
        "PF (Employer)", "ESI (Employer)", "Total Employer Cost",
    ]
    apply_header_row(ws, headers)

    for rec in records:
        emp = rec.employee
        bank = emp.bank_accounts.filter(is_primary=True).first()
        employer_cost = float(rec.gross_earnings + rec.pf_employer + rec.esi_employer + rec.lwf_employer)
        ws.append([
            # Identity
            emp.employee_code,
            emp.full_name,
            emp.department.name if emp.department else "",
            emp.designation.name if emp.designation else "",
            # Bank — full unmasked account for actual disbursement
            bank.bank_name if bank else "",
            bank.account_holder_name if bank else "",
            bank.account_number if bank else "",
            bank.ifsc_code if bank else "",
            bank.branch_name if bank else "",
            bank.account_type if bank else "",
            # Attendance
            rec.working_days,
            float(rec.lop_days),
            float(rec.paid_days),
            # Earnings
            float(rec.basic),
            float(rec.hra),
            float(rec.conveyance),
            float(rec.special_allowance),
            float(rec.other_earnings),
            float(rec.gross_earnings),
            # Statutory deductions
            float(rec.pf_employee),
            float(rec.esi_employee),
            float(rec.professional_tax),
            float(rec.lwf_employee),
            float(rec.tds),
            # Other
            float(rec.loan_deduction),
            float(rec.other_deductions),
            float(rec.total_deductions),
            # Net
            float(rec.net_payable),
            # Employer cost
            float(rec.pf_employer),
            float(rec.esi_employer),
            employer_cost,
        ])

    # Add totals row at the bottom
    last_row = ws.max_row + 1
    from openpyxl.styles import Font, PatternFill
    ws.cell(row=last_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last_row, column=1).fill = PatternFill("solid", fgColor="E8F0FE")
    for col_idx in (14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31):
        col_letter = ws.cell(row=last_row, column=col_idx).column_letter
        cell = ws.cell(row=last_row, column=col_idx, value=f"=SUM({col_letter}2:{col_letter}{last_row-1})")
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8F0FE")

    auto_fit_columns(ws)
    return workbook_response(wb, f"salary_register_{run.year}_{run.month:02d}.xlsx")


@perm_required("payroll.export")
def bank_advice_excel(request, pk):
    """
    Bank advice / disbursement file — India (IFSC) or GCC (IBAN/SWIFT).
    """
    from apps.tenants.jurisdiction import is_gcc_payroll

    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    records = (
        run.records
        .select_related("employee")
        .prefetch_related("employee__bank_accounts")
        .filter(net_payable__gt=0)
        .order_by("employee__employee_code")
    )
    gcc = is_gcc_payroll(tenant.payroll_jurisdiction)
    currency = tenant.currency or ("INR" if not gcc else "AED")

    wb = make_workbook()
    ws = wb.active
    ws.title = f"Bank Advice {run.year}-{run.month:02d}"

    if gcc:
        headers = [
            "Employee Code", "Beneficiary Name", "IBAN", "SWIFT/BIC", "Bank Name",
            f"Amount ({currency})", "Narration",
        ]
    else:
        headers = [
            "Employee Code", "Beneficiary Name", "Account Number", "IFSC", "Bank Name",
            f"Amount ({currency})", "Narration",
        ]
    apply_header_row(ws, headers)

    skipped = []
    for rec in records:
        emp = rec.employee
        bank = emp.bank_accounts.filter(is_primary=True).first()
        if not bank or not bank.account_number:
            skipped.append(emp.employee_code)
            continue
        ws.append([
            emp.employee_code,
            bank.account_holder_name or emp.full_name,
            bank.account_number,
            bank.ifsc_code,
            bank.bank_name,
            float(rec.net_payable),
            f"Salary {run.year}-{run.month:02d} {emp.employee_code}",
        ])

    if skipped:
        ws.append([])
        ws.append([f"WARNING: {len(skipped)} employees skipped (no bank account):"])
        for code in skipped:
            ws.append([code])

    auto_fit_columns(ws)
    return workbook_response(wb, f"bank_advice_{run.year}_{run.month:02d}.xlsx")


@perm_required("payroll.export")
@require_india_payroll
def pf_statement_excel(request, pk):
    """PF ECR format export for EPFO portal upload."""
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    records = run.records.filter(pf_employee__gt=0).select_related("employee")

    wb = make_workbook()
    ws = wb.active
    ws.title = "PF Statement"

    headers = ["UAN", "Member Name", "Gross Wages", "EPF Wages", "EPS Wages",
               "EPF Contribution", "EPS Contribution", "Employer EPF"]
    apply_header_row(ws, headers)

    for rec in records:
        emp = rec.employee
        epf_wages = min(rec.basic, 15000)
        ws.append([
            "",  # UAN — tenant must fill from their EPFO records
            emp.full_name,
            float(rec.gross_earnings),
            float(epf_wages),
            float(epf_wages),
            float(rec.pf_employee),
            float(epf_wages * 833 // 1000),  # EPS = 8.33%
            float(rec.pf_employer),
        ])

    auto_fit_columns(ws)
    return workbook_response(wb, f"pf_ecr_{run.year}_{run.month:02d}.xlsx")


# ---------------------------------------------------------------------------
# ESS: employee payslip
# ---------------------------------------------------------------------------
@login_required
@employee_profile_required
def my_payslips(request):
    employee = request.user.employee_profile
    payslips = Payslip.objects.filter(
        employee=employee, is_published=True
    ).order_by("-year", "-month")
    return render(request, "payroll/my_payslips.html", {"payslips": payslips})


@login_required
def payslip_view(request, pk):
    from django.http import HttpResponse
    from utils.pdf import render_html_to_pdf

    from .payslip_services import render_payslip_html

    employee = getattr(request.user, "employee_profile", None)
    payslip = get_object_or_404(
        Payslip.objects.select_related("payroll_record", "payroll_record__payroll_run", "template"),
        pk=pk,
        employee=employee,
        is_published=True,
    )
    record = payslip.payroll_record
    run = record.payroll_run
    html, _layout = render_payslip_html(record, run, request.tenant, payslip.template)
    pdf_bytes = render_html_to_pdf(html)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="payslip_{payslip.year}_{payslip.month:02d}.pdf"'
    return response


# ---------------------------------------------------------------------------
# Salary structure management
# ---------------------------------------------------------------------------
@perm_required("payroll.configure")
def salary_structure_list(request):
    tenant = request.tenant
    structures = SalaryStructure.objects.filter(tenant=tenant)
    return render(request, "payroll/structures.html", {"structures": structures})


@perm_required("payroll.configure")
@require_india_payroll
def statutory_settings_view(request):
    tenant = request.tenant
    settings = StatutorySetting.objects.filter(tenant=tenant, is_active=True).order_by("statutory_type", "state_code")
    return render(request, "payroll/statutory.html", {"settings": settings})


@perm_required("payroll.configure")
@require_india_payroll
def statutory_create_or_edit(request, pk=None):
    from .forms import StatutorySettingForm
    tenant = request.tenant
    s = get_object_or_404(StatutorySetting, pk=pk, tenant=tenant) if pk else None
    if request.method == "POST":
        form = StatutorySettingForm(request.POST, instance=s)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            obj.slabs = form.cleaned_data.get("slabs_text") or []
            obj.save()
            messages.success(request, "Statutory setting saved.")
            return redirect("payroll:statutory")
    else:
        form = StatutorySettingForm(instance=s)
    return render(request, "payroll/statutory_form.html", {"form": form, "setting": s})


@perm_required("payroll.configure")
def structure_create_or_edit(request, pk=None):
    from .forms import SalaryStructureForm
    tenant = request.tenant
    s = get_object_or_404(SalaryStructure, pk=pk, tenant=tenant) if pk else None
    if request.method == "POST":
        form = SalaryStructureForm(tenant, request.POST, instance=s)
        if form.is_valid():
            form.save()
            messages.success(request, "Salary structure saved.")
            return redirect("payroll:structures")
    else:
        form = SalaryStructureForm(tenant, instance=s)
    return render(request, "payroll/structure_form.html", {"form": form, "structure": s})


# ────────────────────────────────────────────────────────────────────────────
# LOANS — admin & employee views
# ────────────────────────────────────────────────────────────────────────────
@perm_required("payroll.configure")
def loan_list(request):
    """Admin view: all loans across employees."""
    tenant = request.tenant
    status_filter = request.GET.get("status", "active")
    qs = EmployeeLoan.objects.filter(tenant=tenant).select_related("employee")
    if status_filter:
        qs = qs.filter(status=status_filter)
    qs = qs.order_by("-disbursed_date")

    # Stats
    active = EmployeeLoan.objects.filter(tenant=tenant, status="active")
    from django.db.models import Sum
    total_outstanding = active.aggregate(s=Sum("outstanding_amount"))["s"] or 0
    total_principal = active.aggregate(s=Sum("principal_amount"))["s"] or 0

    return render(request, "payroll/loans.html", {
        "loans": qs,
        "status_filter": status_filter,
        "stats": {
            "active_count": active.count(),
            "total_outstanding": total_outstanding,
            "total_principal": total_principal,
        }
    })


@perm_required("payroll.configure")
def loan_create_or_edit(request, pk=None):
    from .forms import EmployeeLoanForm
    tenant = request.tenant
    loan = get_object_or_404(EmployeeLoan, pk=pk, tenant=tenant) if pk else None

    if request.method == "POST":
        form = EmployeeLoanForm(tenant, request.POST, instance=loan)
        if form.is_valid():
            obj = form.save()
            try:
                from apps.hr_ops.services import notify, audit_log
                if obj.employee.user and not pk:
                    notify(
                        obj.employee.user, "general",
                        f"Loan approved: ₹{obj.principal_amount:,.0f}",
                        message=f"A {obj.loan_type or 'loan'} of ₹{obj.principal_amount:,.2f} has been approved. "
                                f"EMI of ₹{obj.emi_amount:,.2f} will be deducted for {obj.total_installments} months "
                                f"starting from the next payroll run.",
                        action_url="/payroll/my-loans/",
                    )
                audit_log(
                    tenant, request.user, "create" if not pk else "update",
                    "EmployeeLoan", obj,
                    f"{'Issued' if not pk else 'Updated'} {obj.loan_type or 'loan'} of ₹{obj.principal_amount} to {obj.employee.full_name}",
                    details={"principal": float(obj.principal_amount), "emi": float(obj.emi_amount)},
                )
            except Exception:
                pass
            messages.success(request, "Loan saved.")
            return redirect("payroll:loans")
    else:
        form = EmployeeLoanForm(tenant, instance=loan)

    return render(request, "payroll/loan_form.html", {"form": form, "loan": loan})


@perm_required("payroll.configure")
def loan_detail(request, pk):
    tenant = request.tenant
    loan = get_object_or_404(EmployeeLoan, pk=pk, tenant=tenant)
    progress = 0
    if loan.total_installments:
        progress = int(loan.paid_installments / loan.total_installments * 100)
    return render(request, "payroll/loan_detail.html", {"loan": loan, "progress": progress})


@login_required
def my_loans(request):
    """Employee view: own loans."""
    employee = getattr(request.user, "employee_profile", None)
    if not employee:
        return redirect("tenants:dashboard")
    loans = EmployeeLoan.objects.filter(employee=employee).order_by("-disbursed_date")
    return render(request, "payroll/my_loans.html", {"loans": loans})


# ────────────────────────────────────────────────────────────────────────────
# REIMBURSEMENTS (Expense Claims) — admin & employee views
# ────────────────────────────────────────────────────────────────────────────
@perm_required("payroll.view_all")
def expense_list(request):
    """Admin view: all expense claims to approve."""
    tenant = request.tenant
    status_filter = request.GET.get("status", "pending")
    qs = ExpenseClaim.objects.filter(tenant=tenant).select_related("employee")
    if status_filter:
        qs = qs.filter(status=status_filter)
    qs = qs.order_by("-expense_date")

    from django.db.models import Sum
    pending = ExpenseClaim.objects.filter(tenant=tenant, status="pending")
    approved_unpaid = ExpenseClaim.objects.filter(tenant=tenant, status="approved", paid_in_run__isnull=True)
    return render(request, "payroll/expenses.html", {
        "expenses": qs,
        "status_filter": status_filter,
        "stats": {
            "pending_count": pending.count(),
            "pending_total": pending.aggregate(s=Sum("amount"))["s"] or 0,
            "to_pay_count": approved_unpaid.count(),
            "to_pay_total": approved_unpaid.aggregate(s=Sum("amount"))["s"] or 0,
        }
    })


@manager_or_hr_required
def team_expenses(request):
    """Manager: pending expense claims from direct reports."""
    tenant = request.tenant
    qs = ExpenseClaim.objects.filter(tenant=tenant, status="pending").select_related("employee")
    if not has_full_team_scope(request.user):
        manager = getattr(request.user, "employee_profile", None)
        if manager:
            qs = qs.filter(employee__reporting_manager=manager)
        else:
            qs = qs.none()
    return render(request, "payroll/team_expenses.html", {
        "expenses": qs.order_by("-expense_date"),
    })


@login_required
@employee_profile_required
def expense_submit(request, pk=None):
    """Employee: submit a new expense claim (or edit a pending one)."""
    from .forms import ExpenseClaimForm
    tenant = request.tenant
    employee = request.user.employee_profile

    claim = None
    if pk:
        claim = get_object_or_404(ExpenseClaim, pk=pk, tenant=tenant, employee=employee, status="pending")

    if request.method == "POST":
        form = ExpenseClaimForm(request.POST, request.FILES, instance=claim)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            obj.employee = employee
            obj.status = "pending"
            obj.save()
            messages.success(request, f"Expense claim of ₹{obj.amount:,.2f} submitted.")

            # Notify HR / reporting manager
            try:
                from apps.hr_ops.services import notify
                approver = employee.reporting_manager.user if employee.reporting_manager and employee.reporting_manager.user else None
                if approver:
                    notify(
                        approver, "general",
                        f"New expense claim from {employee.full_name}",
                        message=f"{employee.full_name} submitted a {obj.category or 'reimbursement'} claim of ₹{obj.amount:,.2f}.\n\n{obj.description}",
                        action_url="/payroll/team-expenses/",
                    )
            except Exception:
                pass
            return redirect("payroll:my_expenses")
    else:
        form = ExpenseClaimForm(instance=claim)

    return render(request, "payroll/expense_form.html", {"form": form, "claim": claim})


@login_required
def my_expenses(request):
    employee = getattr(request.user, "employee_profile", None)
    if not employee:
        return redirect("tenants:dashboard")
    claims = ExpenseClaim.objects.filter(employee=employee).order_by("-created_at")
    return render(request, "payroll/my_expenses.html", {"claims": claims})


@manager_or_hr_required
@require_POST
def expense_action(request, pk):
    """Approve or reject a pending expense claim."""
    tenant = request.tenant
    claim = get_object_or_404(ExpenseClaim, pk=pk, tenant=tenant, status="pending")
    if not can_manage_employee(request.user, claim.employee):
        messages.error(request, "You cannot action this expense claim.")
        next_url = "payroll:team_expenses" if request.user.is_manager and not has_full_team_scope(request.user) else "payroll:expenses"
        return redirect(next_url)
    action = request.POST.get("action")

    if action not in ("approve", "reject"):
        return redirect("payroll:expenses")

    claim.status = "approved" if action == "approve" else "rejected"
    claim.approved_by = request.user
    claim.approved_at = timezone.now()
    claim.save()

    # Notify employee + audit log
    try:
        from apps.hr_ops.services import notify, audit_log
        if claim.employee.user:
            verb = "approved" if action == "approve" else "declined"
            extra = "\n\nIt will be paid out in your next payroll run." if action == "approve" else ""
            notify(
                claim.employee.user, "general",
                f"Your expense claim of ₹{claim.amount:,.0f} was {verb}",
                message=f"{claim.category or 'Reimbursement'} dated {claim.expense_date.strftime('%d %b %Y')}.{extra}",
                action_url="/payroll/my-expenses/",
            )
        audit_log(
            tenant, request.user, "approve" if action == "approve" else "reject",
            "ExpenseClaim", claim,
            f"{'Approved' if action == 'approve' else 'Rejected'} ₹{claim.amount} {claim.category or 'expense'} claim by {claim.employee.full_name}",
            details={"amount": float(claim.amount), "category": claim.category},
        )
    except Exception:
        pass

    messages.success(request, f"Expense {action}d.")
    return redirect("payroll:expenses")


# ---------------------------------------------------------------------------
# Tax declarations (Employee self-service + HR review)
# ---------------------------------------------------------------------------
@login_required
@require_india_payroll
def my_tax_declaration(request):
    """Employee submits/updates their tax declaration for current FY."""
    from .forms import TaxDeclarationForm
    from .tax import current_financial_year, compute_annual_tax
    from decimal import Decimal

    tenant = request.tenant
    emp = getattr(request.user, "employee_profile", None)
    if not emp:
        messages.error(request, "Employee profile not found.")
        return redirect("tenants:dashboard")

    fy = request.GET.get("fy") or current_financial_year()
    declaration, _ = TaxDeclaration.objects.get_or_create(
        tenant=tenant, employee=emp, financial_year=fy,
        defaults={"regime": "new"},
    )

    locked = declaration.status in ("submitted", "verified")

    if request.method == "POST" and not locked:
        form = TaxDeclarationForm(request.POST, instance=declaration)
        action = request.POST.get("action", "save_draft")
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            obj.employee = emp
            obj.financial_year = fy
            if action == "submit":
                obj.status = "submitted"
                obj.submitted_at = timezone.now()
            obj.save()
            try:
                from apps.hr_ops.services import audit_log
                audit_log(
                    tenant, request.user, "submit" if action == "submit" else "update",
                    "TaxDeclaration", obj,
                    f"{'Submitted' if action == 'submit' else 'Saved draft of'} tax declaration "
                    f"for FY{fy} [{obj.regime}]",
                    details={"fy": fy, "regime": obj.regime, "total_80c": float(obj.total_80c)},
                )
            except Exception:
                pass
            messages.success(
                request,
                "Declaration submitted and locked. Contact HR for changes." if action == "submit"
                else "Draft saved. You can keep editing.",
            )
            return redirect("payroll:my_tax_declaration")
    else:
        form = TaxDeclarationForm(instance=declaration)

    # Live tax estimate for the user — pull current salary
    estimate = None
    try:
        salary = EmployeeSalary.objects.filter(
            tenant=tenant, employee=emp, is_active=True,
        ).order_by("-effective_date").first()
        if salary:
            annual_basic = salary.basic_monthly * 12
            annual_ctc = salary.ctc_annual
            overrides = salary.component_overrides or {}
            annual_hra = Decimal(str(overrides.get("HRA", 0))) * 12 or annual_basic * Decimal("0.40")
            estimate = compute_annual_tax(
                regime=declaration.regime,
                gross_salary_annual=annual_ctc,
                basic_annual=annual_basic,
                hra_received_annual=annual_hra,
                declaration=declaration,
            )
    except Exception:
        pass

    return render(request, "payroll/my_tax_declaration.html", {
        "form": form, "declaration": declaration, "fy": fy,
        "locked": locked, "estimate": estimate,
    })


@perm_required("payroll.view_all")
@require_india_payroll
def tax_declaration_admin_list(request):
    """HR view: all employee declarations for an FY."""
    from .tax import current_financial_year
    tenant = request.tenant
    fy = request.GET.get("fy") or current_financial_year()
    declarations = TaxDeclaration.objects.filter(
        tenant=tenant, financial_year=fy,
    ).select_related("employee").order_by("employee__first_name")

    return render(request, "payroll/tax_declarations_admin.html", {
        "declarations": declarations, "fy": fy,
    })


@perm_required("payroll.approve")
@require_POST
@require_india_payroll
def tax_declaration_verify(request, pk):
    """HR marks a declaration as verified."""
    tenant = request.tenant
    decl = get_object_or_404(TaxDeclaration, pk=pk, tenant=tenant)
    decl.status = "verified"
    decl.verified_at = timezone.now()
    decl.verified_by = request.user
    decl.hr_notes = request.POST.get("hr_notes", "")
    decl.save()
    try:
        from apps.hr_ops.services import notify
        if decl.employee.user:
            notify(decl.employee.user, "general",
                   f"Your FY{decl.financial_year} tax declaration was verified",
                   message="Your investment declaration has been reviewed and approved by HR. "
                           "It will be used for monthly TDS computation.",
                   action_url="/payroll/my-tax-declaration/")
    except Exception:
        pass
    messages.success(request, f"Declaration for {decl.employee.full_name} verified.")
    return redirect("payroll:tax_declarations_admin")


# ---------------------------------------------------------------------------
# Form 16 generation (annual TDS certificate)
# ---------------------------------------------------------------------------
@perm_required("payroll.view_all")
@require_india_payroll
def form16_admin_list(request):
    """HR view to bulk-generate Form 16 Part B PDFs for an FY."""
    from .tax import current_financial_year
    tenant = request.tenant
    fy = request.GET.get("fy") or current_financial_year()
    form16s = Form16.objects.filter(tenant=tenant, financial_year=fy).select_related("employee")
    return render(request, "payroll/form16_admin.html", {
        "form16s": form16s, "fy": fy,
    })


@perm_required("payroll.configure")
@require_POST
@require_india_payroll
def form16_generate_all(request):
    """Trigger Form 16 generation for all employees who had payroll in the given FY."""
    from .form16 import generate_form16_for_fy
    tenant = request.tenant
    fy = request.POST.get("fy")
    if not fy:
        messages.error(request, "Financial year is required.")
        return redirect("payroll:form16_admin")

    created, skipped = generate_form16_for_fy(tenant, fy, generated_by=request.user)
    messages.success(request, f"Form 16 Part B generated: {created} created, {skipped} skipped.")
    return redirect("payroll:form16_admin")


@perm_required("payroll.configure")
@require_POST
@require_india_payroll
def form16_issue(request, pk):
    tenant = request.tenant
    form16 = get_object_or_404(Form16, pk=pk, tenant=tenant)
    from .form16_delivery import issue_form16
    if issue_form16(form16, issued_by=request.user):
        messages.success(request, f"Form 16 issued and emailed to {form16.employee.full_name}.")
    else:
        messages.error(request, "Could not issue — ensure Part B PDF exists and employee has email.")
    return redirect(f"{reverse('payroll:form16_admin')}?fy={form16.financial_year}")


@perm_required("payroll.configure")
@require_POST
@require_india_payroll
def form16_issue_all(request):
    tenant = request.tenant
    fy = request.POST.get("fy")
    if not fy:
        messages.error(request, "Financial year is required.")
        return redirect("payroll:form16_admin")
    from .form16_delivery import issue_all_form16_for_fy
    issued, failed = issue_all_form16_for_fy(tenant, fy)
    messages.success(request, f"Issued {issued} Form 16(s). {failed} failed or skipped.")
    return redirect(f"{reverse('payroll:form16_admin')}?fy={fy}")


@login_required
@require_india_payroll
def my_form16s(request):
    """Employee view of their own Form 16s."""
    tenant = request.tenant
    emp = getattr(request.user, "employee_profile", None)
    if not emp:
        return redirect("tenants:dashboard")
    form16s = Form16.objects.filter(tenant=tenant, employee=emp, is_issued=True).order_by("-financial_year")
    return render(request, "payroll/my_form16s.html", {"form16s": form16s})


# ---------------------------------------------------------------------------
# Tally / PF ECR / ESI exports
# ---------------------------------------------------------------------------
@perm_required("payroll.export")
@require_india_payroll
def tally_xml_export(request, pk):
    """Export payroll run as Tally XML voucher batch."""
    from .exports import build_tally_xml
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    xml = build_tally_xml(tenant, run)
    resp = HttpResponse(xml, content_type="application/xml; charset=utf-8")
    resp["Content-Disposition"] = (
        f'attachment; filename="tally_payroll_{run.year}-{run.month:02d}.xml"'
    )
    return resp


@perm_required("payroll.export")
@require_india_payroll
def pf_ecr_export(request, pk):
    """PF ECR file (pipe-delimited TXT per EPFO format)."""
    from .exports import build_pf_ecr
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    txt = build_pf_ecr(tenant, run)
    resp = HttpResponse(txt, content_type="text/plain; charset=utf-8")
    resp["Content-Disposition"] = (
        f'attachment; filename="pf_ecr_{run.year}-{run.month:02d}.txt"'
    )
    return resp


@perm_required("payroll.export")
@require_india_payroll
def esi_return_export(request, pk):
    """ESI monthly contribution return (CSV per ESIC format)."""
    from .exports import build_esi_return
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    csv_text = build_esi_return(tenant, run)
    resp = HttpResponse(csv_text, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = (
        f'attachment; filename="esi_return_{run.year}-{run.month:02d}.csv"'
    )
    return resp


@perm_required("payroll.export")
@require_india_payroll
def statutory_bundle_export(request, pk):
    """ZIP bundle of PF ECR + ESI return for statutory filing."""
    from .exports import build_statutory_zip
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    data = build_statutory_zip(tenant, run)
    resp = HttpResponse(data, content_type="application/zip")
    resp["Content-Disposition"] = (
        f'attachment; filename="statutory_{run.year}-{run.month:02d}.zip"'
    )
    return resp


# ---------------------------------------------------------------------------
# GCC exports (WPS, bank transfer, PIFSS, indemnity liability)
# ---------------------------------------------------------------------------
@perm_required("payroll.export")
@require_gcc_payroll
def wps_sif_export(request, pk):
    from .exports_gcc import build_wps_sif_csv
    from .gcc_export_validation import assess_gcc_export_readiness
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    readiness = assess_gcc_export_readiness(tenant, run)
    if not readiness.can_export:
        messages.error(
            request,
            "Cannot export WPS file — no employees have valid bank IBAN/SWIFT details. "
            "Fix the issues listed on this page first.",
        )
        return redirect("payroll:run_detail", pk=pk)
    if readiness.issue_count:
        messages.warning(
            request,
            f"WPS export includes {readiness.ready_count} employee(s); "
            f"{readiness.issue_count} excluded due to validation issues.",
        )
    csv_text = build_wps_sif_csv(tenant, run)
    resp = HttpResponse(csv_text, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="wps_sif_{run.year}-{run.month:02d}.csv"'
    return resp


@perm_required("payroll.export")
@require_gcc_payroll
def gcc_bank_transfer_export(request, pk):
    from .exports_gcc import build_gcc_bank_transfer_csv
    from .gcc_export_validation import assess_gcc_export_readiness
    tenant = request.tenant
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    readiness = assess_gcc_export_readiness(tenant, run)
    if not readiness.can_export:
        messages.error(
            request,
            "Cannot export bank transfer file — no employees have valid IBAN details.",
        )
        return redirect("payroll:run_detail", pk=pk)
    if readiness.issue_count:
        messages.warning(
            request,
            f"Bank transfer includes {readiness.ready_count} employee(s); "
            f"{readiness.issue_count} excluded due to validation issues.",
        )
    csv_text = build_gcc_bank_transfer_csv(tenant, run)
    resp = HttpResponse(csv_text, content_type="text/csv; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="bank_transfer_{run.year}-{run.month:02d}.csv"'
    return resp


@perm_required("payroll.export")
@require_gcc_payroll
def pifss_statement_export(request, pk):
    """PIFSS contribution report (Kuwait) — Excel."""
    from .exports_gcc import pifss_rows
    tenant = request.tenant
    if normalise_jurisdiction(tenant.payroll_jurisdiction) != "KW":
        messages.info(request, "PIFSS export is for Kuwait workspaces.")
        return redirect("payroll:run_detail", pk=pk)
    run = get_object_or_404(PayrollRun, pk=pk, tenant=tenant)
    wb = make_workbook()
    ws = wb.active
    ws.title = "PIFSS"
    for row in pifss_rows(tenant, run):
        ws.append(row)
    auto_fit_columns(ws)
    return workbook_response(wb, f"pifss_{run.year}_{run.month:02d}.xlsx")


@perm_required("payroll.export")
@require_gcc_payroll
def indemnity_liability_export(request):
    """Active headcount — EOS / indemnity liability snapshot."""
    from .exports_gcc import indemnity_liability_rows
    tenant = request.tenant
    wb = make_workbook()
    ws = wb.active
    ws.title = "EOS Liability"
    for row in indemnity_liability_rows(tenant):
        ws.append(row)
    auto_fit_columns(ws)
    return workbook_response(wb, f"eos_liability_{tenant.subdomain}.xlsx")


# ---------------------------------------------------------------------------
# Payslip templates (per-company layout)
# ---------------------------------------------------------------------------
@perm_required("payroll.configure")
def payslip_template_list(request):
    from apps.hr_ops.letter_company import get_company_profile
    from apps.tenants.jurisdiction import jurisdiction_label
    from .payslip_defaults import DEFAULT_TEMPLATE_NAMES
    from .payslip_services import seed_default_payslip_template

    tenant = request.tenant
    templates = PayslipTemplate.objects.filter(tenant=tenant).order_by("-is_default", "name")
    if not templates.exists():
        seed_default_payslip_template(tenant, created_by=request.user)
        templates = PayslipTemplate.objects.filter(tenant=tenant).order_by("-is_default", "name")
    return render(request, "payroll/payslip_templates.html", {
        "templates": templates,
        "profile": get_company_profile(tenant),
        "layout_labels": dict(PayslipTemplate.LAYOUT_CHOICES),
        "default_names": DEFAULT_TEMPLATE_NAMES,
        "tenant_jurisdiction_label": jurisdiction_label(tenant.payroll_jurisdiction),
    })


@perm_required("payroll.configure")
@require_POST
def payslip_template_seed(request):
    from .payslip_services import seed_default_payslip_template

    created, _skipped = seed_default_payslip_template(request.tenant, created_by=request.user)
    if created:
        messages.success(request, "Installed default payslip template for your region.")
    else:
        messages.info(request, "A payslip template already exists — edit it or add a custom one.")
    return redirect("payroll:payslip_templates")


@perm_required("payroll.configure")
def payslip_template_create_or_edit(request, pk=None):
    from .forms import PayslipTemplateForm
    from .payslip_defaults import default_layout_for_tenant, default_template_name
    from .payslip_services import custom_template_starter

    tenant = request.tenant
    tmpl = get_object_or_404(PayslipTemplate, pk=pk, tenant=tenant) if pk else None
    if request.method == "POST":
        form = PayslipTemplateForm(request.POST, request.FILES, instance=tmpl)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.tenant = tenant
            obj.created_by = obj.created_by or request.user
            obj.save()
            messages.success(request, f'Payslip template "{obj.name}" saved.')
            return redirect("payroll:payslip_templates")
    else:
        initial = {}
        if not tmpl:
            layout = default_layout_for_tenant(tenant)
            initial = {
                "layout": layout,
                "name": default_template_name(layout),
                "template_html": custom_template_starter() if layout == "custom" else "",
                "is_default": True,
            }
        form = PayslipTemplateForm(instance=tmpl, initial=initial if not tmpl else None)
    return render(request, "payroll/payslip_template_form.html", {
        "form": form,
        "template": tmpl,
        "starter_html": custom_template_starter(),
    })


@perm_required("payroll.configure")
def payslip_template_preview(request, pk):
    from decimal import Decimal
    from types import SimpleNamespace
    from utils.pdf import render_html_to_pdf

    from types import SimpleNamespace

    from .payslip_services import render_payslip_html

    class _BankQS:
        def __init__(self, bank):
            self._bank = bank

        def filter(self, **_kwargs):
            return self

        def first(self):
            return self._bank

    tenant = request.tenant
    tmpl = get_object_or_404(PayslipTemplate, pk=pk, tenant=tenant)
    sample_bank = SimpleNamespace(
        bank_name="HDFC Bank",
        account_number="50100123456789",
        is_primary=True,
    )
    sample_record = SimpleNamespace(
        working_days=26,
        paid_days=26,
        lop_days=0,
        basic=Decimal("28000"),
        gross_earnings=Decimal("70000"),
        total_deductions=Decimal("200"),
        net_payable=Decimal("69800"),
        pf_employer=Decimal("0"),
        esi_employer=Decimal("0"),
        earnings_detail={
            "BASIC": {"name": "Basic", "amount": 28000},
            "HRA": {"name": "House Rent Allowance", "amount": 14000},
            "CONV": {"name": "Conveyance", "amount": 2000},
            "OTHER": {"name": "Other Allowance", "amount": 5000},
            "SPECIAL": {"name": "Special Allowance", "amount": 21000},
        },
        deductions_detail={
            "PF_EMP": {"name": "Provident Fund", "amount": 180},
            "PT": {"name": "Professional Tax", "amount": 200},
        },
        employee=SimpleNamespace(
            employee_code="EMP001",
            full_name="Sample Employee",
            name_ar="",
            date_of_joining=datetime.date(2026, 2, 9),
            uan_number="100123456789",
            pan_number="ABCDE1234F",
            department=SimpleNamespace(name="Operations", cost_center_code=""),
            designation=SimpleNamespace(name="Software Engineer", grade=""),
            location=SimpleNamespace(name="Bengaluru"),
            bank_accounts=_BankQS(sample_bank),
        ),
    )
    sample_run = SimpleNamespace(year=2026, month=1)
    html, _ = render_payslip_html(sample_record, sample_run, tenant, tmpl)
    pdf_bytes = render_html_to_pdf(html)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="payslip_preview.pdf"'
    return response
