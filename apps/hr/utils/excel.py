"""
Excel report generation using openpyxl.
"""
import io
from django.http import HttpResponse


def make_workbook():
    """Return a fresh openpyxl Workbook with sensible defaults."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    return wb


def header_style():
    from openpyxl.styles import Font, PatternFill, Alignment
    return {
        "font": Font(bold=True, color="FFFFFF"),
        "fill": PatternFill("solid", fgColor="1F4E79"),
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


def apply_header_row(ws, headers: list, row: int = 1):
    from openpyxl.styles import Font, PatternFill, Alignment
    style = header_style()
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = style["font"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
    ws.row_dimensions[row].height = 20


def workbook_response(wb, filename: str) -> HttpResponse:
    """Stream a Workbook as an HTTP download response."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def auto_fit_columns(ws):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 50)
